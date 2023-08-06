#!/usr/bin/env python
# encoding: utf-8

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from sws_tags.sws_utils import json_encode

def ExportFile(param_export_file):


	queryset = param_export_file['queryset']
	headers = param_export_file['headers']
	col_name = param_export_file['col_name']
	filename = param_export_file['filename']
	logger = param_export_file['logger']


	wb = Workbook()
	ws = wb.get_active_sheet()
	ws.title = filename
	row = 0
	col = 0
	for h in headers:
		cell = ws.cell(row = row, column = col)
		cell.value = h
		col += 1
	row = 1
	col = 0
	for trr in queryset:
		for c in col_name:
			cell = ws.cell(row = row, column = col)
			col += 1
			cell.value = trr[c]
		row += 1
		col = 0
	wb.save("media/excel/{0}".format(filename))
	logger.debug('Excel file saved in media/excel/{0}'.format(filename))
	data =  {
			'msg': '{0} excel file generated, please press download button.'.format(filename),
			'msg_type': 'msg_success',
			'file_url':"excel/{0}".format(filename),
			}
	json_data = json_encode(data)
	return json_data

