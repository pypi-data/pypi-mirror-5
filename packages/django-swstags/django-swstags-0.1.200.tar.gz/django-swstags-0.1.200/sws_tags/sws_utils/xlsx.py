#!/usr/bin/env python
# encoding: utf-8

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from sws_tags.sws_utils import json_encode

import csv
from django.http import HttpResponse
from django.template import loader, Context
from decimal import Decimal
from datetime import date, datetime, timedelta
import time
import ujson
# import datetime

# FOR EXCEL GENERATING
import xlwt
ezxf = xlwt.easyxf

from sws_tags.sws_utils.common_utils import *
from sws_tags.sws_utils.messages import *
from sws_tags.sws_utils.cube import *

import traceback


def ExportFile(param_export_file):


	queryset = param_export_file['queryset']
	headers = param_export_file['headers']
	col_name = param_export_file['col_name']
	filename = param_export_file['filename']
	logger = param_export_file['logger']


	wb = Workbook()
	ws = wb.get_active_sheet()
	ws.title = filename
	row = 0
	col = 0
	for h in headers:
		cell = ws.cell(row = row, column = col)
		cell.value = h
		col += 1
	row = 1
	col = 0
	for trr in queryset:
		for c in col_name:
			cell = ws.cell(row = row, column = col)
			col += 1
			cell.value = trr[c]
		row += 1
		col = 0


	print'aaaaaaaaaaaaaaaaaaaaaaaaa',wb

	wb.save("media/excel/{0}".format(filename))
	logger.debug('Excel file saved in media/excel/{0}'.format(filename))
	data =  {
			'msg': '{0} excel file generated, please press download button.'.format(filename),
			'msg_type': 'msg_success',
			'file_url':"excel/{0}".format(filename),
			}
	json_data = json_encode(data)
	return json_data

def get_csv_query(request, queryset, filename, col_order = None):

	kind_to_xf_map = {
		'date': ezxf(num_format_str='yyyy-mm-dd HH:MM'),
		'int': ezxf(num_format_str='#,##0'),
		'money': ezxf('font: italic on; pattern: pattern solid, fore-colour grey25',
			num_format_str='€#,##0.00'),
		'price': ezxf(num_format_str='#0.000'),
		'text': ezxf(),
		'boolean': ezxf(),
	}

	try:
		# csv_file = open(filename,"w")
		csv_file = HttpResponse()

		csv_writer = csv.writer(csv_file, dialect='excel', quoting=csv.QUOTE_MINIMAL, delimiter='|')

		if type(queryset)!=list:
			q = queryset[0]

			
			fields = []
			for f in queryset._fields:
				fields.append(q[f])
			for f in queryset.aggregate_names:
				fields.append(q[f])

			fields_name = []
			for f in queryset._fields:
				fields_name.append(unicode(getFieldVerboseName(queryset,f)))
			for f in queryset.aggregate_names:
				fields_name.append(f)


			data_xfs_types_used = [k for k in get_field_types(fields)]
			data_xfs = [kind_to_xf_map[k] for k in get_field_types(fields)]

			data = []


			for i, q in enumerate(queryset):
				aux = []
				i=0
				for f in queryset._fields:

					if data_xfs_types_used[i] == 'date':
						date_normalize=request['django_timezone'].normalize(q[f]).strftime('%Y-%m-%d %H:%M:%S')
						aux.append(str(date_normalize))
					else:
						aux.append(q[f])
					i+=1
				for f in queryset.aggregate_names:
					aux.append(q[f])
				data.append(aux)

		else:
			fields_name = []
			if col_order:
				fields_name=col_order
			else:
				for k,v in queryset[0].items():
					fields_name.append(unicode(k))

			data = []

			for q in queryset:
				v_data=[]
				for k in fields_name:
					# print 'excel-->',q[k],'--',k
					v_data.append(q[k])
				data.append(v_data)

		csv_writer.writerow(fields_name)

		for row in data:
			csv_writer.writerow(row)							
								
		# csv_file.close()
		return csv_file

	except Exception, e:
		swslog('error','get_csv_query',e)
		return False


